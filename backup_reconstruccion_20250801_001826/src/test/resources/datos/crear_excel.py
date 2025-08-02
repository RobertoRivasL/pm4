#!/usr/bin/env python3
"""
Script completo para generar TODOS los archivos de datos de prueba
Proyecto: Suite de Automatización Funcional - Roberto Rivas Lopez
Requiere: pip install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import csv
import os
from datetime import datetime

def crear_directorio_datos():
    """Crea el directorio de datos si no existe"""
    if not os.path.exists('datos'):
        os.makedirs('datos')
        print("📁 Directorio 'datos' creado")

def generar_credenciales_csv():
    """Genera el archivo credenciales_ejemplo.csv con datos completos"""
    print("📄 Generando credenciales_ejemplo.csv...")
    
    datos_credenciales = [
        ["usuario", "contrasena", "descripcion", "esperado"],
        ["tomsmith", "SuperSecretPassword!", "Usuario válido por defecto", "true"],
        ["roberto.rivas@test.com", "Password123!", "Usuario Roberto - Proyecto", "true"],
        ["admin@automation.com", "AdminPass456#", "Usuario administrador", "true"],
        ["qa.tester@test.com", "QATest789$", "Usuario de pruebas QA", "true"],
        ["test@email.com", "TestPass123!", "Usuario de prueba genérico", "true"],
        ["usuario.valido@test.com", "SecurePass789!", "Usuario válido adicional", "true"],
        ["test.special@tëst.com", "Spëcïal123!", "Usuario con caracteres especiales", "true"],
        ["usuario.invalido@test.com", "PasswordIncorrecto", "Usuario con password incorrecto", "false"],
        ["usuario.inexistente@fake.com", "Password123!", "Usuario que no existe", "false"],
        ["", "Password123!", "Email vacío", "false"],
        ["roberto.rivas@test.com", "", "Password vacío", "false"],
        ["", "", "Ambos campos vacíos", "false"],
        ["email-sin-arroba.com", "Password123!", "Email con formato inválido", "false"],
        ["usuario con espacios@test.com", "Password123!", "Email con espacios", "false"],
        ["test@test.com", "123", "Password muy corta", "false"],
        ["admin'; DROP TABLE users; --", "Password123!", "Intento de SQL Injection", "false"],
        ["usuario@test.com", "<script>alert('xss')</script>", "Intento de XSS en password", "false"],
        ["very.long.email@very-long-domain.com", "VeryLongPassword123!", "Credenciales largas", "false"],
        ["usuario@test", "Password123!", "Email sin dominio completo", "false"],
        ["@test.com", "Password123!", "Email sin parte local", "false"]
    ]
    
    with open('credenciales_ejemplo.csv', 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerows(datos_credenciales)
    
    print(f"✅ credenciales_ejemplo.csv creado con {len(datos_credenciales)-1} casos de prueba")

def generar_excel_completo():
    """Genera el archivo usuarios_ejemplo.xlsx completo"""
    print("📊 Generando usuarios_ejemplo.xlsx...")
    
    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # === HOJA 1: DATOS VÁLIDOS ===
    ws1 = wb.create_sheet("Datos_Registro_Validos")
    
    encabezados1 = [
        "caso_prueba", "descripcion", "nombre", "apellido", "email", "password", 
        "confirmar_password", "telefono", "genero", "pais", "ciudad", 
        "fecha_nacimiento", "es_valido", "resultado_esperado"
    ]
    
    # Agregar encabezados con estilo
    for col, header in enumerate(encabezados1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos válidos
    datos_validos = [
        ["REG_001", "Usuario completo - Roberto Rivas", "Roberto", "Rivas", 
         "roberto.rivas.test@email.com", "Password123!", "Password123!", "+56912345678", 
         "Masculino", "Chile", "Santiago", "15/03/1995", True, "Registro exitoso"],
        
        ["REG_002", "Usuario QA - Ana García", "Ana", "García", 
         "ana.garcia.qa@test.com", "SecurePass456#", "SecurePass456#", "+56987654321", 
         "Femenino", "Chile", "Valparaíso", "22/08/1987", True, "Registro exitoso"],
        
        ["REG_003", "Usuario con caracteres especiales", "José María", "Pérez-González", 
         "jose.maria@test.com", "StrongPwd789$", "StrongPwd789$", "+34987654321", 
         "Masculino", "España", "Madrid", "10/12/1980", True, "Registro exitoso"],
        
        ["REG_004", "Usuario internacional", "John Michael", "Smith-Johnson", 
         "john.international@global.com", "MySecure2024!", "MySecure2024!", "+1234567890", 
         "Masculino", "Estados Unidos", "Nueva York", "05/07/1990", True, "Registro exitoso"],
        
        ["REG_005", "Usuario con email complejo", "María Elena", "Rodríguez-López", 
         "maria.elena+test@sub-domain.co.uk", "TestPass123#", "TestPass123#", "+44123456789", 
         "Femenino", "Reino Unido", "Londres", "18/11/1985", True, "Registro exitoso"],
        
        ["REG_006", "Usuario de automatización", "Carlos", "Mendoza", 
         "carlos.automation@test.com", "AutoTest456$", "AutoTest456$", "+56976543210", 
         "Masculino", "Chile", "Concepción", "30/01/1992", True, "Registro exitoso"],
        
        ["REG_007", "Usuario con apellido compuesto", "Isabella", "Fernández-Torres", 
         "isabella.compound@test.org", "Secure789!", "Secure789!", "+34912345678", 
         "Femenino", "España", "Barcelona", "14/04/1988", True, "Registro exitoso"],
        
        ["REG_008", "Usuario con caracteres latinos", "André", "Müller", 
         "andre.latin@test.de", "LatinChars123#", "LatinChars123#", "+49123456789", 
         "Masculino", "Alemania", "Berlín", "25/09/1983", True, "Registro exitoso"]
    ]
    
    for row, data in enumerate(datos_validos, 2):
        for col, value in enumerate(data, 1):
            ws1.cell(row=row, column=col, value=value)
    
    # === HOJA 2: DATOS INVÁLIDOS ===
    ws2 = wb.create_sheet("Datos_Registro_Invalidos")
    
    encabezados2 = [
        "caso_prueba", "descripcion", "nombre", "apellido", "email", "password", 
        "confirmar_password", "telefono", "es_valido", "mensaje_error"
    ]
    
    for col, header in enumerate(encabezados2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        cell.alignment = header_alignment
    
    datos_invalidos = [
        ["REG_INV_001", "Nombre vacío", "", "García", 
         "nombre.vacio@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "El nombre es obligatorio"],
        
        ["REG_INV_002", "Apellido vacío", "Roberto", "", 
         "apellido.vacio@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "El apellido es obligatorio"],
        
        ["REG_INV_003", "Email sin @", "Roberto", "Rivas", 
         "email-sin-arroba.com", "Password123!", "Password123!", "+56912345678", 
         False, "Formato de email inválido"],
        
        ["REG_INV_004", "Contraseñas diferentes", "Roberto", "Rivas", 
         "passwords.diferentes@test.com", "Password123!", "DiferentePass456!", "+56912345678", 
         False, "Las contraseñas no coinciden"],
        
        ["REG_INV_005", "Contraseña muy corta", "Roberto", "Rivas", 
         "password.corta@test.com", "123", "123", "+56912345678", 
         False, "La contraseña debe tener al menos 8 caracteres"],
        
        ["REG_INV_006", "Campos obligatorios vacíos", "", "", 
         "", "", "", "", 
         False, "Todos los campos obligatorios son requeridos"],
        
        ["REG_INV_007", "Email con espacios", "Roberto", "Rivas", 
         "email con espacios@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "El email no puede contener espacios"],
        
        ["REG_INV_008", "Inyección XSS", "<script>alert('XSS')</script>", "Rivas", 
         "xss.test@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "Caracteres no permitidos en el nombre"],
        
        ["REG_INV_009", "Nombre muy largo", "A" * 51, "Apellido", 
         "nombre.largo@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "El nombre excede la longitud máxima"],
        
        ["REG_INV_010", "Teléfono inválido", "Roberto", "Rivas", 
         "telefono.invalido@test.com", "Password123!", "Password123!", "telefono-abc-123", 
         False, "Formato de teléfono inválido"],
        
        ["REG_INV_011", "SQL Injection", "Roberto'; DROP TABLE users; --", "Apellido", 
         "sql.injection@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "Caracteres no permitidos"],
        
        ["REG_INV_012", "Password sin mayúsculas", "Roberto", "Rivas", 
         "password.minuscula@test.com", "password123!", "password123!", "+56912345678", 
         False, "La contraseña debe contener mayúsculas"],
        
        ["REG_INV_013", "Password sin números", "Roberto", "Rivas", 
         "password.sin.numeros@test.com", "PasswordSinNumeros!", "PasswordSinNumeros!", "+56912345678", 
         False, "La contraseña debe contener números"],
        
        ["REG_INV_014", "Email muy largo", "Roberto", "Rivas", 
         "usuario.con.email.extremadamente.largo@dominio-muy-largo-que-excede-limites.com", "Password123!", "Password123!", "+56912345678", 
         False, "Email excede longitud máxima"],
        
        ["REG_INV_015", "Caracteres HTML en email", "Roberto", "Rivas", 
         "<b>email</b>@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "Caracteres HTML no permitidos"]
    ]
    
    for row, data in enumerate(datos_invalidos, 2):
        for col, value in enumerate(data, 1):
            ws2.cell(row=row, column=col, value=value)
    
    # === HOJA 3: DATOS LOGIN ===
    ws3 = wb.create_sheet("Datos_Login_Mixtos")
    
    encabezados3 = [
        "caso_prueba", "descripcion", "email", "password", "es_valido", 
        "resultado_esperado", "mensaje_error"
    ]
    
    for col, header in enumerate(encabezados3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = header_alignment
    
    datos_login = [
        ["LOGIN_XLSX_001", "Login exitoso - Roberto", "roberto.rivas.test@email.com", 
         "Password123!", True, "Login exitoso y redirección", ""],
        
        ["LOGIN_XLSX_002", "Login exitoso - Admin", "admin.automation@test.com", 
         "AdminPass456#", True, "Login con permisos administrativos", ""],
        
        ["LOGIN_XLSX_003", "Login exitoso - QA", "qa.automation@test.com", 
         "QAPass789$", True, "Login para usuario de pruebas", ""],
        
        ["LOGIN_XLSX_004", "Login fallido - Usuario inexistente", "usuario.inexistente@test.com", 
         "Password123!", False, "Login rechazado", "Usuario no encontrado"],
        
        ["LOGIN_XLSX_005", "Login fallido - Password incorrecta", "roberto.rivas.test@email.com", 
         "PasswordIncorrecto", False, "Login rechazado", "Contraseña incorrecta"],
        
        ["LOGIN_XLSX_006", "Login fallido - Email malformado", "email.malformado@", 
         "Password123!", False, "Error de validación", "Formato de email inválido"],
        
        ["LOGIN_XLSX_007", "Login con caracteres especiales", "test.special@tëst.com", 
         "Spëcïal123!", True, "Login exitoso", ""],
        
        ["LOGIN_XLSX_008", "Login fallido - SQL Injection", "admin'; DROP TABLE users; --", 
         "Password123!", False, "Error de seguridad", "Caracteres no permitidos"],
        
        ["LOGIN_XLSX_009", "Login fallido - XSS", "usuario.test@test.com", 
         "<script>alert('xss')</script>", False, "Error de seguridad", "Caracteres no permitidos"],
        
        ["LOGIN_XLSX_010", "Login fallido - Email muy largo", "usuario.extremadamente.largo@dominio-muy-largo.com", 
         "Password123!", False, "Error de validación", "Email excede longitud máxima"]
    ]
    
    for row, data in enumerate(datos_login, 2):
        for col, value in enumerate(data, 1):
            ws3.cell(row=row, column=col, value=value)
    
    # Aplicar formato a todas las hojas
    for ws in [ws1, ws2, ws3]:
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 3, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Aplicar bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
    
    # Guardar archivo
    wb.save("usuarios_ejemplo.xlsx")
    print(f"✅ usuarios_ejemplo.xlsx creado con {len(datos_validos)} casos válidos, {len(datos_invalidos)} inválidos y {len(datos_login)} de login")

def generar_usuarios_adicionales_csv():
    """Genera el archivo usuarios_adicionales.csv"""
    print("📄 Generando usuarios_adicionales.csv...")
    
    datos_adicionales = [
        ["caso_prueba", "descripcion", "nombre", "apellido", "email", "password", "confirmar_password", "telefono", "genero", "pais", "ciudad", "fecha_nacimiento", "es_valido", "resultado_esperado", "mensaje_error"],
        ["REG_ADD_001", "Usuario adicional 1", "Pedro", "Martinez", "pedro.martinez@test.com", "Password123!", "Password123!", "+56912345678", "Masculino", "Chile", "Santiago", "01/01/1990", "true", "Registro exitoso", ""],
        ["REG_ADD_002", "Usuario adicional 2", "Laura", "Silva", "laura.silva@test.com", "SecurePass456#", "SecurePass456#", "+56987654321", "Femenino", "Chile", "Valparaíso", "15/05/1985", "true", "Registro exitoso", ""],
        ["REG_ADD_003", "Usuario adicional 3", "Miguel", "Torres", "miguel.torres@test.com", "StrongPwd789$", "StrongPwd789$", "+56976543210", "Masculino", "Chile", "Concepción", "20/12/1988", "true", "Registro exitoso", ""],
        ["REG_ADD_004", "Error - Email duplicado", "Roberto", "Rivas", "roberto.rivas.test@email.com", "Password123!", "Password123!", "+56912345678", "", "", "", "", "false", "Error de registro", "Email ya existe"],
        ["REG_ADD_005", "Error - Contraseña muy simple", "Usuario", "Prueba", "usuario.prueba@test.com", "123456", "123456", "+56912345678", "", "", "", "", "false", "Error de validación", "Contraseña muy simple"]
    ]
    
    with open('usuarios_adicionales.csv', 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerows(datos_adicionales)
    
    print(f"✅ usuarios_adicionales.csv creado con {len(datos_adicionales)-1} casos")

def main():
    """Función principal que genera todos los archivos"""
    print("🚀 GENERADOR COMPLETO DE DATOS DE PRUEBA")
    print("Proyecto: Suite de Automatización Funcional")
    print("Autor: Roberto Rivas Lopez")
    print("=" * 50)
    
    try:
        crear_directorio_datos()
        generar_credenciales_csv()
        generar_excel_completo()
        generar_usuarios_adicionales_csv()
        
        print("\n" + "=" * 50)
        print("✅ TODOS LOS ARCHIVOS GENERADOS EXITOSAMENTE")
        print("📁 Archivos creados:")
        print("   • credenciales_ejemplo.csv")
        print("   • usuarios_ejemplo.xlsx")
        print("   • usuarios_adicionales.csv")
        print("\n🎯 Archivos listos para usar en el proyecto de automatización")
        
    except ImportError:
        print("❌ ERROR: Se requiere instalar openpyxl")
        print("💡 Ejecuta: pip install openpyxl")
    except Exception as e:
        print(f"❌ ERROR: {e}")

if __name__ == "__main__":
    main()