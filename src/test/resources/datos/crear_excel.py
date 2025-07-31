#!/usr/bin/env python3
"""
Script corregido para generar el archivo usuarios_prueba.xlsx
con los datos específicos del proyecto de Roberto Rivas Lopez
Requiere: pip install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

def crear_excel_proyecto():
    print("🚀 Generando archivo usuarios_prueba.xlsx para el proyecto...")
    
    # Crear workbook
    wb = openpyxl.Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # === HOJA 1: DATOS VÁLIDOS DE REGISTRO ===
    print("📄 Creando hoja: Datos_Registro_Validos...")
    ws1 = wb.create_sheet("Datos_Registro_Validos")
    
    # Encabezados para datos válidos
    encabezados1 = [
        "caso_prueba", "descripcion", "nombre", "apellido", "email", "password", 
        "confirmar_password", "telefono", "genero", "pais", "ciudad", 
        "fecha_nacimiento", "es_valido", "resultado_esperado"
    ]
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Agregar encabezados
    for col, header in enumerate(encabezados1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos válidos reales del proyecto
    datos_validos_reales = [
        ["REG_001", "Usuario completo válido - Roberto Rivas", "Roberto", "Rivas", 
         "roberto.rivas.test@email.com", "Password123!", "Password123!", "+56912345678", 
         "Masculino", "Chile", "Santiago", "15/03/1995", True, "Registro exitoso"],
        
        ["REG_002", "Usuario con datos mínimos", "Ana", "García", 
         "ana.garcia.qa@test.com", "SecurePass456#", "SecurePass456#", "", 
         "Femenino", "Chile", "Valparaíso", "22/08/1987", True, "Registro exitoso"],
        
        ["REG_003", "Usuario con caracteres especiales", "José María", "Pérez-González", 
         "jose.maria.especial@test.com", "StrongPwd789$", "StrongPwd789$", "+56987654321", 
         "Masculino", "España", "Madrid", "10/12/1980", True, "Registro exitoso"],
        
        ["REG_004", "Cliente internacional", "John Michael", "Smith-Johnson", 
         "john.international@global.com", "MySecure2024!", "MySecure2024!", "+1234567890", 
         "Masculino", "Estados Unidos", "Nueva York", "05/07/1990", True, "Registro exitoso"],
        
        ["REG_005", "Usuario con email complejo", "María Elena", "Rodríguez-López", 
         "maria.elena+automation@sub-domain.co.uk", "TestPass123#", "TestPass123#", "+44123456789", 
         "Femenino", "Reino Unido", "Londres", "18/11/1985", True, "Registro exitoso"],
        
        ["REG_006", "Usuario de QA Testing", "Carlos", "Mendoza", 
         "carlos.qa.tester@automation.com", "QATester456$", "QATester456$", "+56976543210", 
         "Masculino", "Chile", "Concepción", "30/01/1992", True, "Registro exitoso"],
        
        ["REG_007", "Usuaria con apellido compuesto", "Isabella", "Fernández-Torres", 
         "isabella.compound@test.org", "Secure789!", "Secure789!", "+34987654321", 
         "Femenino", "España", "Barcelona", "14/04/1988", True, "Registro exitoso"],
        
        ["REG_008", "Usuario con caracteres latinos", "André", "Müller", 
         "andre.latin@tëst.com", "LatinChars123#", "LatinChars123#", "+49123456789", 
         "Masculino", "Alemania", "Berlín", "25/09/1983", True, "Registro exitoso"],
    ]
    
    # Agregar datos válidos
    for row, data in enumerate(datos_validos_reales, 2):
        for col, value in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=value)
            if col == 13:  # Columna es_valido
                cell.value = bool(value)
    
    # === HOJA 2: DATOS INVÁLIDOS DE REGISTRO ===
    print("📄 Creando hoja: Datos_Registro_Invalidos...")
    ws2 = wb.create_sheet("Datos_Registro_Invalidos")
    
    # Encabezados para datos inválidos
    encabezados2 = [
        "caso_prueba", "descripcion", "nombre", "apellido", "email", "password", 
        "confirmar_password", "telefono", "es_valido", "mensaje_error"
    ]
    
    # Agregar encabezados
    for col, header in enumerate(encabezados2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
        cell.alignment = header_alignment
    
    # Datos inválidos reales del proyecto
    datos_invalidos_reales = [
        ["REG_INV_001", "Validación: Nombre vacío", "", "García", 
         "nombre.vacio@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "El nombre es obligatorio"],
        
        ["REG_INV_002", "Validación: Apellido vacío", "Roberto", "", 
         "apellido.vacio@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "El apellido es obligatorio"],
        
        ["REG_INV_003", "Validación: Email sin @", "Roberto", "Rivas", 
         "email-sin-arroba.com", "Password123!", "Password123!", "+56912345678", 
         False, "Formato de email inválido"],
        
        ["REG_INV_004", "Validación: Email sin dominio", "Roberto", "Rivas", 
         "email@", "Password123!", "Password123!", "+56912345678", 
         False, "Formato de email inválido"],
        
        ["REG_INV_005", "Validación: Contraseñas diferentes", "Roberto", "Rivas", 
         "passwords.diferentes@test.com", "Password123!", "DiferentePass456!", "+56912345678", 
         False, "Las contraseñas no coinciden"],
        
        ["REG_INV_006", "Validación: Contraseña muy corta", "Roberto", "Rivas", 
         "password.corta@test.com", "123", "123", "+56912345678", 
         False, "La contraseña debe tener al menos 8 caracteres"],
        
        ["REG_INV_007", "Validación: Contraseña sin mayúsculas", "Roberto", "Rivas", 
         "password.minusculas@test.com", "password123!", "password123!", "+56912345678", 
         False, "La contraseña debe contener al menos una mayúscula"],
        
        ["REG_INV_008", "Validación: Contraseña sin números", "Roberto", "Rivas", 
         "password.sin.numeros@test.com", "PasswordSinNumeros!", "PasswordSinNumeros!", "+56912345678", 
         False, "La contraseña debe contener al menos un número"],
        
        ["REG_INV_009", "Validación: Email con espacios", "Roberto", "Rivas", 
         "email con espacios@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "El email no puede contener espacios"],
        
        ["REG_INV_010", "Security: Inyección XSS en nombre", "<script>alert('XSS')</script>", "Rivas", 
         "xss.injection@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "Caracteres no permitidos en el nombre"],
        
        ["REG_INV_011", "Security: SQL Injection en apellido", "Roberto', DROP TABLE users; --", "Rivas", 
         "sql.injection@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "Caracteres no permitidos en el apellido"],
        
        ["REG_INV_012", "Validación: Campos obligatorios vacíos", "", "", 
         "", "", "", "", 
         False, "Todos los campos obligatorios deben ser completados"],
        
        ["REG_INV_013", "Validación: Nombre excesivamente largo", "A" * 51, "Apellido", 
         "nombre.largo@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "El nombre no puede exceder 50 caracteres"],
        
        ["REG_INV_014", "Validación: Teléfono con formato inválido", "Roberto", "Rivas", 
         "telefono.invalido@test.com", "Password123!", "Password123!", "telefono-123-abc", 
         False, "Formato de teléfono inválido"],
        
        ["REG_INV_015", "Security: HTML en email", "Roberto", "Rivas", 
         "<b>email</b>@test.com", "Password123!", "Password123!", "+56912345678", 
         False, "Caracteres HTML no permitidos en email"],
    ]
    
    # Agregar datos inválidos
    for row, data in enumerate(datos_invalidos_reales, 2):
        for col, value in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            if col == 9:  # Columna es_valido
                cell.value = bool(value)
    
    # === HOJA 3: DATOS LOGIN MIXTOS ===
    print("📄 Creando hoja: Datos_Login_Mixtos...")
    ws3 = wb.create_sheet("Datos_Login_Mixtos")
    
    # Encabezados para login
    encabezados3 = [
        "caso_prueba", "descripcion", "email", "password", "es_valido", 
        "resultado_esperado", "mensaje_error"
    ]
    
    # Agregar encabezados
    for col, header in enumerate(encabezados3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = header_alignment
    
    # Datos login específicos del proyecto
    datos_login_proyecto = [
        ["LOGIN_XLSX_001", "Login exitoso - Usuario Roberto", "roberto.rivas.test@email.com", 
         "Password123!", True, "Login exitoso y redirección al dashboard", ""],
        
        ["LOGIN_XLSX_002", "Login exitoso - Usuario Admin", "admin.proyecto@test.com", 
         "AdminPass456#", True, "Login exitoso con permisos de administrador", ""],
        
        ["LOGIN_XLSX_003", "Login exitoso - Usuario QA", "qa.automation@test.com", 
         "QAPass789$", True, "Login exitoso para usuario de pruebas", ""],
        
        ["LOGIN_XLSX_004", "Login fallido - Email inexistente", "usuario.inexistente@test.com", 
         "Password123!", False, "Login rechazado", "Usuario no encontrado en el sistema"],
        
        ["LOGIN_XLSX_005", "Login fallido - Contraseña incorrecta", "roberto.rivas.test@email.com", 
         "PasswordIncorrecto", False, "Login rechazado", "Contraseña incorrecta"],
        
        ["LOGIN_XLSX_006", "Login fallido - Email malformado", "email.malformado@", 
         "Password123!", False, "Error de validación", "Formato de email inválido"],
        
        ["LOGIN_XLSX_007", "Login con caracteres especiales", "test.special@tëst.com", 
         "Spëcïal123!", True, "Login exitoso con caracteres especiales", ""],
        
        ["LOGIN_XLSX_008", "Login fallido - Inyección SQL", "admin'; DROP TABLE users; --", 
         "Password123!", False, "Error de seguridad", "Caracteres no permitidos"],
        
        ["LOGIN_XLSX_009", "Login fallido - XSS en password", "usuario.test@test.com", 
         "<script>alert('xss')</script>", False, "Error de seguridad", "Caracteres no permitidos"],
        
        ["LOGIN_XLSX_010", "Login con email muy largo", "usuario.con.email.extremadamente.largo.que.podria.causar.problemas@dominio-muy-largo.com", 
         "Password123!", False, "Error de validación", "Email excede longitud máxima"],
    ]
    
    # Agregar datos login
    for row, data in enumerate(datos_login_proyecto, 2):
        for col, value in enumerate(data, 1):
            cell = ws3.cell(row=row, column=col, value=value)
            if col == 5:  # Columna es_valido
                cell.value = bool(value)
    
    # === AJUSTAR FORMATO DE TODAS LAS HOJAS ===
    print("🎨 Aplicando formato a las hojas...")
    for ws in [ws1, ws2, ws3]:
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            # Importante: para openpyxl 3.0+ 'column' es un iterador de celdas, no un objeto de columna directamente
            # Necesitamos el letter de la primera celda para la dimensión de columna
            column_letter = column[0].column_letter 
            
            for cell in column:
                try:
                    # Sumar 1 para dar un pequeño margen
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError): # Manejar posibles errores con tipos de datos no convertibles a string
                    pass
            
            # Limitar el ancho máximo para evitar columnas excesivamente grandes
            adjusted_width = min(max_length + 3, 60)  # Máximo 60 caracteres, más 3 de padding
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Aplicar bordes a todas las celdas con datos
        # Ya se importó Border y Side al principio, así que no es necesario importarlo aquí de nuevo
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Iterar sobre las celdas que contienen datos (desde la fila 1 hasta la última con datos)
        # Esto es más eficiente que iterar sobre todo el rango potencial de filas y columnas.
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None: # Aplicar borde solo si la celda tiene valor
                    cell.border = thin_border
    
    # === GUARDAR ARCHIVO ===
    filename = "usuarios_prueba.xlsx"
    wb.save(filename)
    
    print(f"✅ Archivo '{filename}' creado exitosamente!")
    print(f"📊 Resumen de datos generados:")
    print(f"    • Hoja 1: {len(datos_validos_reales)} casos de registro válidos")
    print(f"    • Hoja 2: {len(datos_invalidos_reales)} casos de registro inválidos") 
    print(f"    • Hoja 3: {len(datos_login_proyecto)} casos de login mixtos")
    print(f"    • Total: {len(datos_validos_reales) + len(datos_invalidos_reales) + len(datos_login_proyecto)} casos de prueba")
    print(f"📍 Ubicación: {os.path.abspath(filename)}")

if __name__ == "__main__":
    try:
        crear_excel_proyecto()
    except ImportError:
        print("❌ Error: Necesitas instalar openpyxl")
        print("💡 Ejecuta: pip install openpyxl")
    except Exception as e:
        print(f"❌ Error al crear el archivo: {e}")